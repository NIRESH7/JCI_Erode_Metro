"""Generate JCI member bulk-import Excel template matching admin panel field labels."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = r"c:\Users\Admin\Desktop\jci\JCI_Member_Import_Template.xlsx"

# Exact labels from admin panel Create Member form (top to bottom).
COLUMNS = [
    "Profile Picture",
    "Username",
    "Membership ID",
    "Email",
    "Contact",
    "Gender",
    "DOB",
    "Communication Address",
    "Blood Group",
    "Willing to Donate",
    "Company Name",
    "Business Category",
    "Designation",
    "Board Member",
    "Marital Status",
    "Role",
    "JCI location",
]

INSTRUCTIONS = [
    ("Admin panel field", "Description", "Allowed values / format"),
    ("Profile Picture", "Photo file name (upload separately in admin panel)", "Optional — e.g. photo.jpg"),
    ("Username", "Full name of the member", "Required"),
    ("Membership ID", "JCI membership number", "Text or number"),
    ("Email", "Email address", "Required, must be unique"),
    ("Contact", "Mobile number", "10 digits only (e.g. 9876543210). +91 is added automatically."),
    ("Gender", "Gender", "male | female | others"),
    ("DOB", "Date of birth", "dd-mm-yyyy (e.g. 16-04-2001)"),
    ("Communication Address", "Full address", "Free text"),
    ("Blood Group", "Blood group", "O+, O-, A+, A-, B+, B-, AB+, AB-, A1+, A2+, A1B+, A1B-, A2B+, HH"),
    ("Willing to Donate", "Willing to donate blood", "yes | no"),
    ("Company Name", "Company / office name", "Free text"),
    ("Business Category", "Business category", "As shown in admin dropdown"),
    ("Designation", "Job designation", "As shown in admin dropdown"),
    ("Board Member", "Is this a board member?", "yes | no"),
    ("Marital Status", "Marital status", "As shown in admin dropdown"),
    ("Role", "JCI role", "As shown in admin dropdown (e.g. Member, President)"),
    ("JCI location", "JCI chapter / location", "As shown in admin dropdown"),
    ("", "", ""),
    ("Note", "Fill one member per row. Do not change the header row.", ""),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(wrap_text=True, vertical="center")

    for col_idx, label in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(label) + 2)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    last_row = 51

    validations = [
        ("F", '"male,female,others"'),       # Gender
        ("I", '"O+,O-,A+,A-,B+,B-,AB+,AB-,A1+,A2+,A1B+,A1B-,A2B+,HH"'),  # Blood Group
        ("J", '"yes,no"'),                   # Willing to Donate
        ("N", '"yes,no"'),                   # Board Member
    ]
    for col_letter, formula in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please choose a value from the dropdown list."
        dv.errorTitle = "Invalid value"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")

    info = wb.create_sheet("Instructions")
    info["A1"] = "JCI Member Import Template"
    info["A1"].font = Font(bold=True, size=14)
    info.merge_cells("A1:C1")

    info["A3"] = (
        "Column headers match the admin panel Create Member form exactly. "
        "Fill one member per row on the Members sheet."
    )
    info.merge_cells("A3:C3")

    start = 5
    for r, row in enumerate(INSTRUCTIONS, start=start):
        for c, value in enumerate(row, start=1):
            cell = info.cell(row=r, column=c, value=value)
            if r == start:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")

    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 48
    info.column_dimensions["C"].width = 52

    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
