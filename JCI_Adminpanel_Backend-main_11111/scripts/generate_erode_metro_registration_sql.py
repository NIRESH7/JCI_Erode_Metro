"""
Generate add_members_erode_metro_registration.sql from the Google Form
registration Excel + numbered profile images (1.*, 2.*, ...).

Usage:
  py generate_erode_metro_registration_sql.py
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
METRO_ROOT = SCRIPT_DIR.parent.parent
DEFAULT_INPUT = METRO_ROOT / "JCI Erode Metro Member Registration Form (Web &APP).xlsx"
DEFAULT_IMAGES = METRO_ROOT / "images"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "sql database" / "add_members_erode_metro_registration.sql"
DEFAULT_DATABASE = "api_jcierodemetro"
JCI_LOCATION = "JCI Erode Metro"
PLACEHOLDER_PIC = "/images/placeholder.jpg"

# Extra member not in Excel (requested separately).
EXTRA_MEMBERS: list[dict[str, str | None]] = [
    {
        "profile_pic": PLACEHOLDER_PIC,
        "user_name": "Niresh",
        "membership_id": None,
        "email": "nireshtoranto@gmail.com",
        "contact": "+919345034653",
        "gender": "male",
        "dob": "15/06/1995",
        "location": "Perundurai Road, Erode, Tamil Nadu",
        "blood_group": "B+",
        "willing_to_donate": "yes",
        "office_name": "Nutz Technovation",
        "sector": "Information Technology",
        "job": "Software Developer",
        "member_type": "member",
        "board_member": "no",
        "martial_status": "unmarried",
        "role": "Member",
        "jci_location": JCI_LOCATION,
    },
]

# Column indices on "Form Responses 1" (0-based)
COL = {
    "profile_drive": 1,
    "user_name": 2,
    "membership_id": 3,
    "role": 4,
    "email": 6,
    "contact": 7,
    "gender": 8,
    "dob": 9,
    "location": 10,
    "blood_group": 11,
    "willing": 12,
    "marital": 13,
    "job": 23,
    "office_name": 24,
    "sector": 25,
}

VALID_BLOOD_GROUPS = {
    "O+", "O-", "A+", "A-", "B+", "B-", "AB+", "AB-",
    "A1+", "A2+", "A1B+", "A1B-", "A2B+", "HH",
}


def sql_escape(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def normalize_contact(raw: str) -> str:
    digits = re.sub(r"\D", "", raw or "")
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) != 10:
        raise ValueError(f"Contact must be 10 digits, got {raw!r}")
    return "+91" + digits


def normalize_dob(raw) -> str | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%d/%m/%Y")
    text = str(raw).strip().rstrip("-").strip()
    for fmt in (
        "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%m-%d-%Y",
        "%Y-%m-%d", "%d-%m-%y", "%d/%m/%y", "%m/%d/%y",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    raise ValueError(f"DOB invalid: {raw!r}")


def normalize_blood_group(raw: str) -> str | None:
    if not raw:
        return None
    text = re.sub(r"\s+", "", raw.strip().upper())
    if text not in VALID_BLOOD_GROUPS:
        raise ValueError(f"Blood Group invalid: {raw!r}")
    return text


def normalize_willing(raw: str) -> str | None:
    value = (raw or "").strip().lower()
    if value == "yes":
        return "yes"
    if value == "no":
        return "no"
    return None  # Maybe / empty


def normalize_marital(raw: str) -> str | None:
    value = (raw or "").strip().lower()
    if value in ("married", "unmarried"):
        return value
    return None


def normalize_gender(raw: str) -> str:
    value = (raw or "").strip().lower()
    if value in ("male", "female", "others"):
        return value
    return "male"


def normalize_role_type(raw: str) -> tuple[str, str, str]:
    """Returns (role_label, member_type, board_flag)."""
    role = (raw or "Member").strip()
    lower = role.lower()
    if lower in ("board member", "past president"):
        return role, "boardmember", "yes"
    return role or "Member", "member", "no"


def image_path_for_index(images_dir: Path, index: int) -> str:
    matches = list(images_dir.glob(f"{index}.*"))
    if not matches:
        raise FileNotFoundError(f"No image for index {index} in {images_dir}")
    # Prefer exact stem match (avoid 1 vs 10)
    exact = [p for p in matches if p.stem == str(index)]
    if not exact:
        raise FileNotFoundError(f"No exact image stem {index} in {images_dir}")
    return f"/images/{exact[0].name}"


def read_rows(path: Path) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Form Responses 1"
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet {sheet_name!r} not found. Sheets: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows[1:]  # skip header


def transform(row, index: int, images_dir: Path) -> dict[str, str | None]:
    def get(key: str):
        i = COL[key]
        return row[i] if i < len(row) else None

    user_name = cell_str(get("user_name"))
    email = cell_str(get("email")).lower()
    contact = normalize_contact(cell_str(get("contact")))
    membership_id = cell_str(get("membership_id")) or None
    role, member_type, board_flag = normalize_role_type(cell_str(get("role")))

    if not user_name:
        raise ValueError(f"Row {index}: Username required")
    if not email or "@" not in email:
        raise ValueError(f"Row {index}: Email invalid: {email!r}")

    return {
        "profile_pic": image_path_for_index(images_dir, index),
        "user_name": user_name,
        "membership_id": membership_id,
        "email": email,
        "contact": contact,
        "gender": normalize_gender(cell_str(get("gender"))),
        "dob": normalize_dob(get("dob")),
        "location": cell_str(get("location")) or None,
        "blood_group": normalize_blood_group(cell_str(get("blood_group"))),
        "willing_to_donate": normalize_willing(cell_str(get("willing"))),
        "office_name": cell_str(get("office_name")) or None,
        "sector": cell_str(get("sector")) or None,
        "job": cell_str(get("job")) or None,
        "member_type": member_type,
        "board_member": board_flag,
        "martial_status": normalize_marital(cell_str(get("marital"))),
        "role": role,
        "jci_location": JCI_LOCATION,
    }


def build_sql(rows: list[dict[str, str | None]], database: str, source_file: str) -> str:
    lines: list[str] = [
        "-- =============================================================================",
        "-- JCI Erode Metro — member import from registration form Excel",
        f"-- Source: {source_file}",
        f"-- Members: {len(rows)} (Excel + extras)",
        "-- Profile pics: /images/1.* … /images/N.* (row order); extras may use placeholder",
        "--",
        "-- Before running: backup the live database.",
        "-- Also upload images 1.*–N.* to the API static /images folder.",
        "--",
        "-- Usage:",
        f"--   mysql -u jcierodemetro -p {database} < add_members_erode_metro_registration.sql",
        "--   Or phpMyAdmin -> Import -> this file",
        "-- =============================================================================",
        "",
        f"USE `{database}`;",
        "",
        "SET NAMES utf8mb4;",
        "SET FOREIGN_KEY_CHECKS = 0;",
        "",
        "START TRANSACTION;",
        "",
        "DROP TEMPORARY TABLE IF EXISTS `MemberImportStaging`;",
        "",
        "CREATE TEMPORARY TABLE `MemberImportStaging` (",
        "  `profile_pic` varchar(255) NOT NULL,",
        "  `user_name` varchar(255) NOT NULL,",
        "  `membership_id` varchar(255) DEFAULT NULL,",
        "  `email` varchar(255) NOT NULL,",
        "  `contact` varchar(255) NOT NULL,",
        "  `gender` enum('male','female','others') NOT NULL DEFAULT 'male',",
        "  `dob` varchar(15) DEFAULT NULL,",
        "  `location` text DEFAULT NULL,",
        "  `blood_group` enum('O+','O-','A+','A-','B+','B-','AB+','AB-','A1+','A2+','A1B+','A1B-','A2B+','HH') DEFAULT NULL,",
        "  `willing_to_donate` enum('yes','no') DEFAULT NULL,",
        "  `office_name` varchar(255) DEFAULT NULL,",
        "  `sector` varchar(255) DEFAULT NULL,",
        "  `job` varchar(255) DEFAULT NULL,",
        "  `member_type` enum('member','boardmember') NOT NULL DEFAULT 'member',",
        "  `board_member` enum('yes','no') NOT NULL DEFAULT 'no',",
        "  `martial_status` varchar(255) DEFAULT NULL,",
        "  `role` varchar(255) DEFAULT NULL,",
        "  `jci_location` varchar(255) DEFAULT NULL",
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;",
        "",
        "INSERT INTO `MemberImportStaging` (",
        "  `profile_pic`, `user_name`, `membership_id`, `email`, `contact`,",
        "  `gender`, `dob`, `location`, `blood_group`, `willing_to_donate`,",
        "  `office_name`, `sector`, `job`, `member_type`, `board_member`,",
        "  `martial_status`, `role`, `jci_location`",
        ") VALUES",
    ]

    value_lines = []
    for row in rows:
        value_lines.append(
            "("
            f"{sql_escape(row['profile_pic'])}, "
            f"{sql_escape(row['user_name'])}, "
            f"{sql_escape(row['membership_id'])}, "
            f"{sql_escape(row['email'])}, "
            f"{sql_escape(row['contact'])}, "
            f"{sql_escape(row['gender'])}, "
            f"{sql_escape(row['dob'])}, "
            f"{sql_escape(row['location'])}, "
            f"{sql_escape(row['blood_group'])}, "
            f"{sql_escape(row['willing_to_donate'])}, "
            f"{sql_escape(row['office_name'])}, "
            f"{sql_escape(row['sector'])}, "
            f"{sql_escape(row['job'])}, "
            f"{sql_escape(row['member_type'])}, "
            f"{sql_escape(row['board_member'])}, "
            f"{sql_escape(row['martial_status'])}, "
            f"{sql_escape(row['role'])}, "
            f"{sql_escape(row['jci_location'])}"
            ")"
        )

    lines.append(",\n".join(value_lines) + ";")
    lines.extend([
        "",
        "INSERT INTO `Member` (",
        "  `profile_pic`, `user_name`, `membership_id`, `email`, `contact`,",
        "  `gender`, `dob`, `location`, `blood_group`, `willing_to_donate`,",
        "  `office_name`, `job`, `sector`, `martial_status`, `role`, `jci_location`,",
        "  `type`, `status`, `app_access`, `createdAt`, `updatedAt`",
        ")",
        "SELECT",
        "  s.`profile_pic`, s.`user_name`, s.`membership_id`, s.`email`, s.`contact`,",
        "  s.`gender`, s.`dob`, s.`location`, s.`blood_group`, s.`willing_to_donate`,",
        "  s.`office_name`, s.`job`, s.`sector`, s.`martial_status`, s.`role`, s.`jci_location`,",
        "  s.`member_type`, 'active', 'view', NOW(), NOW()",
        "FROM `MemberImportStaging` s",
        "WHERE NOT EXISTS (",
        "  SELECT 1 FROM `Member` m",
        "  WHERE m.`email` = s.`email` OR m.`contact` = s.`contact`",
        ");",
        "",
        "INSERT INTO `boardMembers` (`member_id`, `createdAt`, `updatedAt`)",
        "SELECT m.`id`, NOW(), NOW()",
        "FROM `Member` m",
        "INNER JOIN `MemberImportStaging` s ON s.`email` = m.`email`",
        "WHERE s.`board_member` = 'yes'",
        "  AND NOT EXISTS (",
        "    SELECT 1 FROM `boardMembers` b WHERE b.`member_id` = m.`id`",
        "  );",
        "",
        "DROP TEMPORARY TABLE IF EXISTS `MemberImportStaging`;",
        "",
        "SET FOREIGN_KEY_CHECKS = 1;",
        "",
        "COMMIT;",
        "",
        "SELECT 'Members imported (total active)' AS check_type, COUNT(*) AS value",
        "FROM `Member` WHERE `status` = 'active';",
        "SELECT 'Board members linked' AS check_type, COUNT(*) AS value FROM `boardMembers`;",
        "",
    ])
    return "\n".join(lines)


def main() -> int:
    input_path = DEFAULT_INPUT
    images_dir = DEFAULT_IMAGES
    output_path = DEFAULT_OUTPUT

    if not input_path.exists():
        print(f"Error: Excel not found: {input_path}", file=sys.stderr)
        return 1
    if not images_dir.is_dir():
        print(f"Error: Images folder not found: {images_dir}", file=sys.stderr)
        return 1

    raw_rows = read_rows(input_path)
    transformed: list[dict[str, str | None]] = []
    errors: list[str] = []
    seen_emails: set[str] = set()
    seen_contacts: set[str] = set()

    member_index = 0
    for row in raw_rows:
        # Skip empty rows (no name / email)
        name = cell_str(row[COL["user_name"]]) if len(row) > COL["user_name"] else ""
        email = cell_str(row[COL["email"]]) if len(row) > COL["email"] else ""
        if not name and not email:
            continue
        member_index += 1
        try:
            data = transform(row, member_index, images_dir)
        except Exception as exc:
            errors.append(str(exc))
            continue

        if data["email"] in seen_emails:
            errors.append(f"Duplicate email in Excel: {data['email']}")
            continue
        if data["contact"] in seen_contacts:
            errors.append(f"Duplicate contact in Excel: {data['contact']}")
            continue
        seen_emails.add(data["email"] or "")
        seen_contacts.add(data["contact"] or "")
        transformed.append(data)

    for extra in EXTRA_MEMBERS:
        email = extra["email"] or ""
        contact = extra["contact"] or ""
        if email not in seen_emails and contact not in seen_contacts:
            transformed.append(extra)
            seen_emails.add(email)
            seen_contacts.add(contact)

    if errors:
        print("Validation issues:", file=sys.stderr)
        for err in errors:
            print(f"  - {err}", file=sys.stderr)

    if not transformed:
        print("No valid rows to import.", file=sys.stderr)
        return 1

    sql = build_sql(transformed, DEFAULT_DATABASE, input_path.name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(sql, encoding="utf-8")

    board_count = sum(1 for r in transformed if r["board_member"] == "yes")
    print(f"Read {member_index} member row(s) from {input_path.name}")
    print(f"Generated {len(transformed)} INSERT(s), boardmembers={board_count}")
    for r in transformed:
        print(f"  {r['profile_pic']:20} {r['user_name'][:40]:40} {r['member_type']}")
    print(f"Wrote: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
