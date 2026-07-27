"""
Read JCI_Member_Import_Template.xlsx and generate member_bulk_import.sql.

Usage:
  py generate_member_import_sql.py
  py generate_member_import_sql.py --input "../../JCI_Member_Import_Template (1).xlsx"
  py generate_member_import_sql.py --database api_jcierodegreencity
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR.parent.parent / "JCI_Member_Import_Template (1).xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "sql database" / "member_bulk_import.sql"
DEFAULT_DATABASE = "api_jcierodegreencity"

EXCEL_COLUMNS = [
    "Profile Picture",
    "Username",
    "Membership ID",
    "Email",
    "Contact",
    "Gender",
    "DOB",
    "Communication Address",
    "Blood Group",
    "Willing to Donate",
    "Company Name",
    "Business Category",
    "Designation",
    "Board Member",
    "Marital Status",
    "Role",
    "JCI location",
]

VALID_GENDERS = {"male", "female", "others"}
VALID_BLOOD_GROUPS = {
    "O+", "O-", "A+", "A-", "B+", "B-", "AB+", "AB-",
    "A1+", "A2+", "A1B+", "A1B-", "A2B+", "HH",
}
VALID_YES_NO = {"yes", "no"}
VALID_MARITAL = {"married", "unmarried"}
PLACEHOLDER_PIC = "/images/placeholder.jpg"

# Extra member not in Excel (requested separately).
EXTRA_MEMBERS: list[dict[str, str | None]] = [
    {
        "profile_pic": PLACEHOLDER_PIC,
        "user_name": "Niresh Toranto",
        "membership_id": None,
        "email": "nireshtoranto@gmail.com",
        "contact": "+919345034653",
        "gender": "male",
        "dob": "15/06/1995",
        "location": "Perundurai Road, Erode, Tamil Nadu",
        "blood_group": "B+",
        "willing_to_donate": "yes",
        "office_name": "Nutz Technovation",
        "sector": "Information Technology",
        "job": "Software Developer",
        "member_type": "member",
        "board_member": "no",
        "martial_status": "unmarried",
        "role": "Member",
        "jci_location": "JCI Erode Greencity",
    },
]


def sql_escape(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    return str(value).strip()


def is_empty_row(values: dict[str, str]) -> bool:
    key_fields = ("Username", "Email", "Contact")
    return not any(values.get(k) for k in key_fields)


def normalize_contact(raw: str) -> str:
    digits = re.sub(r"\D", "", raw or "")
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) != 10:
        raise ValueError(f"Contact must be 10 digits, got {raw!r}")
    return "+91" + digits


def normalize_dob(raw: str) -> str:
    if not raw:
        return ""
    text = raw.strip().rstrip("-").strip()
    for fmt in (
        "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%m-%d-%Y",
        "%Y-%m-%d", "%d-%m-%y", "%d/%m/%y", "%m/%d/%y",
    ):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%d/%m/%Y")
        except ValueError:
            continue
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%d/%m/%Y")
    raise ValueError(f"DOB invalid (use dd-mm-yyyy): {raw!r}")


def normalize_profile_pic(raw: str) -> str:
    if not raw:
        return PLACEHOLDER_PIC
    text = raw.strip()
    if text.startswith("http://") or text.startswith("https://"):
        return text
    if text.startswith("/images/"):
        return text
    filename = Path(text).name
    return f"/images/{filename}"


def normalize_board_type(raw: str) -> tuple[str, str]:
    value = (raw or "no").strip().lower()
    if value not in VALID_YES_NO:
        raise ValueError(f"Board Member must be yes or no, got {raw!r}")
    return ("boardmember" if value == "yes" else "member", value)


def normalize_enum(raw: str, allowed: set[str], field: str, optional: bool = True) -> str | None:
    if not raw:
        return None if optional else ""
    value = raw.strip().lower()
    if value not in allowed:
        raise ValueError(f"{field} must be one of {sorted(allowed)}, got {raw!r}")
    return value


def normalize_blood_group(raw: str) -> str | None:
    if not raw:
        return None
    text = re.sub(r"\s+", " ", raw.strip().lower())
    compact = re.sub(r"\s+", "", text)

    aliases = {
        "o+ve": "O+", "o+positive": "O+", "o+postive": "O+", "o+ve": "O+",
        "o +ve": "O+", "o positive": "O+",
        "o-ve": "O-", "o negative": "O-",
        "a+ve": "A+", "a +ve": "A+", "a positive": "A+",
        "a-ve": "A-", "a negative": "A-",
        "b+ve": "B+", "b +ve": "B+", "b positive": "B+",
        "b-ve": "B-", "b negative": "B-",
        "ab+ve": "AB+", "ab +ve": "AB+", "ab positive": "AB+",
        "ab-ve": "AB-", "ab negative": "AB-",
        "a1+ve": "A1+", "a1 postive": "A1+", "a1 positive": "A1+",
        "a1-ve": "A1-", "a1 negative": "A1-",
        "a2+ve": "A2+", "a2 positive": "A2+",
        "a1b+ve": "A1B+", "a1b positive": "A1B+",
        "a1b-ve": "A1B-", "a1b negative": "A1B-",
        "a2b+ve": "A2B+", "a2b positive": "A2B+",
        "hh": "HH",
    }

    value = aliases.get(text, aliases.get(compact, raw.strip().upper().replace(" ", "")))
    mapping = {"A1B+": "A1B+", "A1B-": "A1B-", "A2B+": "A2B+"}
    value = mapping.get(value, value)
    if value not in VALID_BLOOD_GROUPS:
        raise ValueError(f"Blood Group invalid: {raw!r}")
    return value


def read_members_sheet(path: Path) -> list[tuple[int, dict[str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Members" not in wb.sheetnames:
        raise SystemExit(f"Sheet 'Members' not found in {path}")
    ws = wb["Members"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    headers = [cell_str(h) for h in rows[0]]
    if headers[: len(EXCEL_COLUMNS)] != EXCEL_COLUMNS:
        raise SystemExit(
            "Excel header row does not match template.\n"
            f"Expected: {EXCEL_COLUMNS}\n"
            f"Found:    {headers[: len(EXCEL_COLUMNS)]}"
        )

    members: list[tuple[int, dict[str, str]]] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        values = {
            EXCEL_COLUMNS[i]: cell_str(row[i]) if i < len(row) else ""
            for i in range(len(EXCEL_COLUMNS))
        }
        if is_empty_row(values):
            continue
        members.append((row_idx, values))
    return members


def transform_row(row_num: int, raw: dict[str, str]) -> dict[str, str | None]:
    errors: list[str] = []

    user_name = raw["Username"].strip()
    email = raw["Email"].strip().lower()
    contact_raw = raw["Contact"].strip()

    if not user_name:
        errors.append("Username is required")
    if not email:
        errors.append("Email is required")
    elif "@" not in email:
        errors.append(f"Email invalid: {email!r}")

    try:
        contact = normalize_contact(contact_raw) if contact_raw else ""
        if not contact_raw:
            errors.append("Contact is required")
    except ValueError as exc:
        errors.append(str(exc))
        contact = ""

    try:
        dob = normalize_dob(raw["DOB"]) if raw["DOB"] else None
    except ValueError as exc:
        errors.append(str(exc))
        dob = None

    try:
        gender = normalize_enum(raw["Gender"], VALID_GENDERS, "Gender")
    except ValueError as exc:
        errors.append(str(exc))
        gender = None

    try:
        blood_group = normalize_blood_group(raw["Blood Group"])
    except ValueError as exc:
        errors.append(str(exc))
        blood_group = None

    try:
        willing = normalize_enum(raw["Willing to Donate"], VALID_YES_NO, "Willing to Donate")
    except ValueError as exc:
        errors.append(str(exc))
        willing = None

    try:
        marital = normalize_enum(raw["Marital Status"], VALID_MARITAL, "Marital Status")
    except ValueError as exc:
        errors.append(str(exc))
        marital = None

    try:
        member_type, board_flag = normalize_board_type(raw["Board Member"])
    except ValueError as exc:
        errors.append(str(exc))
        member_type, board_flag = "member", "no"

    if errors:
        raise ValueError(f"Row {row_num}: " + "; ".join(errors))

    return {
        "profile_pic": normalize_profile_pic(raw["Profile Picture"]),
        "user_name": user_name,
        "membership_id": raw["Membership ID"].strip() or None,
        "email": email,
        "contact": contact,
        "gender": gender or "male",
        "dob": dob,
        "location": raw["Communication Address"].strip() or None,
        "blood_group": blood_group,
        "willing_to_donate": willing,
        "office_name": raw["Company Name"].strip() or None,
        "sector": raw["Business Category"].strip() or None,
        "job": raw["Designation"].strip() or None,
        "member_type": member_type,
        "board_member": board_flag,
        "martial_status": marital,
        "role": raw["Role"].strip() or None,
        "jci_location": raw["JCI location"].strip() or None,
    }


def build_sql(rows: list[dict[str, str | None]], database: str, source_file: str) -> str:
    lines: list[str] = [
        "-- =============================================================================",
        "-- JCI Member bulk import (generated from Excel)",
        f"-- Source: {source_file}",
        f"-- Members: {len(rows)} (includes extra member: Niresh Toranto)",
        "--",
        "-- Before running:",
        "--   1. Run jci_production_live_setup.sql first (fresh schema).",
        "--   2. Backup the database.",
        "--",
        "-- Usage:",
        f"--   mysql -u jcierodegreencity -p {database} < member_bulk_import.sql",
        "--   Or phpMyAdmin -> Import -> this file",
        "-- =============================================================================",
        "",
        f"USE `{database}`;",
        "",
        "SET NAMES utf8mb4;",
        "SET FOREIGN_KEY_CHECKS = 0;",
        "",
        "START TRANSACTION;",
        "",
        "DROP TEMPORARY TABLE IF EXISTS `MemberImportStaging`;",
        "",
        "CREATE TEMPORARY TABLE `MemberImportStaging` (",
        "  `profile_pic` varchar(255) NOT NULL,",
        "  `user_name` varchar(255) NOT NULL,",
        "  `membership_id` varchar(255) DEFAULT NULL,",
        "  `email` varchar(255) NOT NULL,",
        "  `contact` varchar(255) NOT NULL,",
        "  `gender` enum('male','female','others') NOT NULL DEFAULT 'male',",
        "  `dob` varchar(15) DEFAULT NULL,",
        "  `location` text DEFAULT NULL,",
        "  `blood_group` enum('O+','O-','A+','A-','B+','B-','AB+','AB-','A1+','A2+','A1B+','A1B-','A2B+','HH') DEFAULT NULL,",
        "  `willing_to_donate` enum('yes','no') DEFAULT NULL,",
        "  `office_name` varchar(255) DEFAULT NULL,",
        "  `sector` varchar(255) DEFAULT NULL,",
        "  `job` varchar(255) DEFAULT NULL,",
        "  `member_type` enum('member','boardmember') NOT NULL DEFAULT 'member',",
        "  `board_member` enum('yes','no') NOT NULL DEFAULT 'no',",
        "  `martial_status` varchar(255) DEFAULT NULL,",
        "  `role` varchar(255) DEFAULT NULL,",
        "  `jci_location` varchar(255) DEFAULT NULL",
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;",
        "",
        "INSERT INTO `MemberImportStaging` (",
        "  `profile_pic`, `user_name`, `membership_id`, `email`, `contact`,",
        "  `gender`, `dob`, `location`, `blood_group`, `willing_to_donate`,",
        "  `office_name`, `sector`, `job`, `member_type`, `board_member`,",
        "  `martial_status`, `role`, `jci_location`",
        ") VALUES",
    ]

    value_lines = []
    for row in rows:
        value_lines.append(
            "("
            f"{sql_escape(row['profile_pic'])}, "
            f"{sql_escape(row['user_name'])}, "
            f"{sql_escape(row['membership_id'])}, "
            f"{sql_escape(row['email'])}, "
            f"{sql_escape(row['contact'])}, "
            f"{sql_escape(row['gender'])}, "
            f"{sql_escape(row['dob'])}, "
            f"{sql_escape(row['location'])}, "
            f"{sql_escape(row['blood_group'])}, "
            f"{sql_escape(row['willing_to_donate'])}, "
            f"{sql_escape(row['office_name'])}, "
            f"{sql_escape(row['sector'])}, "
            f"{sql_escape(row['job'])}, "
            f"{sql_escape(row['member_type'])}, "
            f"{sql_escape(row['board_member'])}, "
            f"{sql_escape(row['martial_status'])}, "
            f"{sql_escape(row['role'])}, "
            f"{sql_escape(row['jci_location'])}"
            ")"
        )

    lines.append(",\n".join(value_lines) + ";")
    lines.extend([
        "",
        "INSERT INTO `Member` (",
        "  `profile_pic`, `user_name`, `membership_id`, `email`, `contact`,",
        "  `gender`, `dob`, `location`, `blood_group`, `willing_to_donate`,",
        "  `office_name`, `job`, `sector`, `martial_status`, `role`, `jci_location`,",
        "  `type`, `status`, `app_access`, `createdAt`, `updatedAt`",
        ")",
        "SELECT",
        "  s.`profile_pic`, s.`user_name`, s.`membership_id`, s.`email`, s.`contact`,",
        "  s.`gender`, s.`dob`, s.`location`, s.`blood_group`, s.`willing_to_donate`,",
        "  s.`office_name`, s.`job`, s.`sector`, s.`martial_status`, s.`role`, s.`jci_location`,",
        "  s.`member_type`, 'active', 'view', NOW(), NOW()",
        "FROM `MemberImportStaging` s",
        "WHERE NOT EXISTS (",
        "  SELECT 1 FROM `Member` m",
        "  WHERE m.`email` = s.`email` OR m.`contact` = s.`contact`",
        ");",
        "",
        "INSERT INTO `boardMembers` (`member_id`, `createdAt`, `updatedAt`)",
        "SELECT m.`id`, NOW(), NOW()",
        "FROM `Member` m",
        "INNER JOIN `MemberImportStaging` s ON s.`email` = m.`email`",
        "WHERE s.`board_member` = 'yes'",
        "  AND NOT EXISTS (",
        "    SELECT 1 FROM `boardMembers` b WHERE b.`member_id` = m.`id`",
        "  );",
        "",
        "DROP TEMPORARY TABLE IF EXISTS `MemberImportStaging`;",
        "",
        "SET FOREIGN_KEY_CHECKS = 1;",
        "",
        "COMMIT;",
        "",
        "SELECT 'Members imported (total active)' AS check_type, COUNT(*) AS value",
        "FROM `Member` WHERE `status` = 'active';",
        "SELECT 'Board members linked' AS check_type, COUNT(*) AS value FROM `boardMembers`;",
    ])
    return "\n".join(lines) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate member_bulk_import.sql from Excel")
    parser.add_argument("--input", "-i", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", "-o", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--database", "-d", default=DEFAULT_DATABASE)
    parser.add_argument("--strict", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    if not input_path.exists():
        print(f"Error: Excel file not found: {input_path}", file=sys.stderr)
        return 1

    raw_members = read_members_sheet(input_path)
    transformed: list[dict[str, str | None]] = []
    errors: list[str] = []
    seen_emails: set[str] = set()
    seen_contacts: set[str] = set()

    for row_num, raw in raw_members:
        try:
            row = transform_row(row_num, raw)
        except ValueError as exc:
            errors.append(str(exc))
            if args.strict:
                continue
            continue

        email = row["email"] or ""
        contact = row["contact"] or ""
        if email in seen_emails:
            errors.append(f"Row {row_num}: duplicate email in Excel: {email}")
            continue
        if contact in seen_contacts:
            errors.append(f"Row {row_num}: duplicate contact in Excel: {contact}")
            continue
        seen_emails.add(email)
        seen_contacts.add(contact)
        transformed.append(row)

    for extra in EXTRA_MEMBERS:
        email = extra["email"] or ""
        contact = extra["contact"] or ""
        if email not in seen_emails and contact not in seen_contacts:
            transformed.append(extra)
            seen_emails.add(email)
            seen_contacts.add(contact)

    if not transformed:
        print("No valid rows to import.", file=sys.stderr)
        return 1

    if errors:
        print("Validation issues (some rows skipped):", file=sys.stderr)
        for err in errors:
            print(f"  - {err}", file=sys.stderr)

    if args.strict and errors:
        return 1

    sql = build_sql(transformed, args.database, input_path.name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(sql, encoding="utf-8")

    print(f"Read {len(raw_members)} row(s) from {input_path}")
    print(f"Generated {len(transformed)} member INSERT(s)")
    print(f"Wrote: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
